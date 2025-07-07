import requests
import pytz
import os
from dotenv import load_dotenv
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from collections import defaultdict

load_dotenv()
API_KEY = os.getenv("API_KEY")
API_URL = 'https://api.backupradar.com/backups'

headers = {
    'accept': 'application/json',
    'ApiKey': API_KEY
}

# Get yesterday's date in AEST
aest = pytz.timezone('Australia/Sydney')
yesterday_aest = datetime.now(aest) - timedelta(days=1)
target_date_str = yesterday_aest.strftime('%Y-%m-%dT00:00:00')

# Pagination setup
page = 1
page_size = 1000
all_results = []

statuses_to_include = ['Failure', 'Warning', 'No Result', 'Pending']
special_keywords = [
    'OneDrive to Cloud storage',
    'Office 365 mailboxes to Cloud storage',
    'SharePoint sites to Cloud storage',
    'Microsoft 365 mailboxes to Cloud storage',
    'Microsoft Teams to Cloud storage'
]

# Fetch data
while True:
    params = {
        'Page': page,
        'PageSize': page_size,
        'date': target_date_str,
        'statuses': statuses_to_include,
        'FilterScheduled': 'true'
    }

    response = requests.get(API_URL, headers=headers, params=params)
    if response.status_code != 200:
        print(f"[ERROR] Failed to retrieve data on page {page}: {response.status_code} - {response.text}")
        break

    data = response.json()
    results = data.get('Results', [])
    if not results:
        break

    all_results.extend(results)

    total_pages = data.get('TotalPages', page)
    if page >= total_pages:
        break
    page += 1

# Separate jobs
regular_jobs = []
special_acronis_jobs = []

for backup in all_results:
    method = backup.get('methodName', '')
    job_name = backup.get('jobName', '')
    if method in ['Acronis API', 'Acronis'] and any(keyword in job_name for keyword in special_keywords):
        special_acronis_jobs.append(backup)
    else:
        regular_jobs.append(backup)

# Sort
def sort_key(backup):
    return (backup.get('companyName', ''), backup.get('methodName', ''))

regular_jobs.sort(key=sort_key)
special_acronis_jobs.sort(key=sort_key)

# Excel formatting
wb = Workbook()
ws = wb.active
ws.title = "Backup Report"

bold_font = Font(bold=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
client_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

# M365 section styling
m365_title_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
m365_header_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
m365_client_fill = PatternFill(start_color="C5E0B4", end_color="C5E0B4", fill_type="solid")

# Conditional formatting fills
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

fieldnames = [
    'Server/Workload Affected',
    'Status',
    'Job Name',
    'Backup Method',
    'Resolved',
    'Ticket number',
    'Technician Notes'
]

# Group jobs by client
def group_by_client(jobs):
    grouped = defaultdict(list)
    for job in jobs:
        client = job.get('companyName', 'Unknown Client')
        grouped[client].append(job)
    return grouped

# Add checkbox-style data validation
checkbox_validation = DataValidation(type="list", formula1='"✘,☑"', allow_blank=True)
ws.add_data_validation(checkbox_validation)

# Write grouped jobs
def write_jobs(jobs, client_fill, header_fill, row_num, is_m365=False):
    for client, jobs in group_by_client(jobs).items():
        client_cell = ws.cell(row=row_num, column=1, value=client)
        client_cell.font = bold_font
        client_cell.fill = client_fill
        client_cell.border = thin_border
        ws.row_dimensions[row_num].height = 20
        row_num += 1

        for col_num, header in enumerate(fieldnames, 1):
            cell = ws.cell(row=row_num, column=col_num, value=header)
            cell.font = bold_font
            cell.border = thin_border
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row_num].height = 20
        row_num += 1

        start_row = row_num
        for i, job in enumerate(jobs):
            row = [
                job.get('deviceName', ''),
                job.get('status', {}).get('name', ''),
                job.get('jobName', ''),
                job.get('methodName', ''),
                '✘',  # Default to unresolved
                '',   # Empty ticket number
                ''    # Empty technician notes
            ]
            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border

                # Wrap text in Technician Notes column
                if col_num == 6:
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

                if col_num == 5:
                    checkbox_validation.add(cell)
                if i % 2 == 1:
                    cell.fill = zebra_fill
            ws.row_dimensions[row_num].height = 30  # Slightly taller for wrapped notes
            row_num += 1
        end_row = row_num - 1

        # Apply conditional formatting
        col_letter = get_column_letter(5)
        for r in range(start_row, end_row + 1):
            ws.conditional_formatting.add(f"{col_letter}{r}", FormulaRule(formula=[f"${col_letter}{r}=\"☑\""], fill=green_fill))
            ws.conditional_formatting.add(f"{col_letter}{r}", FormulaRule(formula=[f"${col_letter}{r}=\"✘\""], fill=red_fill))

        row_num += 1
    return row_num

row_num = 1
row_num = write_jobs(regular_jobs, client_fill, header_fill, row_num)

# Add spacing and styled label for M365 section
row_num += 1
m365_title_cell = ws.cell(row=row_num, column=1, value="M365 Acronis Backups")
m365_title_cell.font = Font(bold=True)
m365_title_cell.fill = m365_title_fill
m365_title_cell.border = thin_border
ws.row_dimensions[row_num].height = 20
row_num += 1

row_num = write_jobs(special_acronis_jobs, m365_client_fill, m365_header_fill, row_num, is_m365=True)

# Auto-size columns
for col in ws.columns:
    max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
    col_letter = get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

# Ensure Technician Notes column is wide enough
notes_col_letter = get_column_letter(7)
ws.column_dimensions[notes_col_letter].width = max(ws.column_dimensions[notes_col_letter].width, 100)

# Save file
excel_file = f'enhanced_backup_report_{yesterday_aest.date()}_AEST.xlsx'
wb.save(excel_file)

# Terminal summary
print(f"[OK] Formatted Excel report saved as '{excel_file}'")
print(f"[INFO] Total records included: {len(regular_jobs) + len(special_acronis_jobs)}")
print(f"[INFO] Regular jobs: {len(regular_jobs)}")
print(f"[INFO] Special Acronis jobs: {len(special_acronis_jobs)}")
